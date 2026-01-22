"""账户管理服务"""

from datetime import datetime
from nanoid import generate
from decimal import Decimal
from io import BytesIO
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from app.models import UserAccount, RechargeRecord, TokenUsageRecord
from app.database import db_session, db_transaction
from app.utils import get_logger
from app.config import Config

logger = get_logger(__name__)


def get_or_create_account(openid: str):
    """获取或创建用户账户"""
    with db_transaction() as db:
        account = (
            db.query(UserAccount)
            .filter(UserAccount.user_openid == openid, UserAccount.is_deleted == False)
            .first()
        )

        if not account:
            balance = 0.00
            if openid in Config.SEED_USERS:
                balance = 1000000
            account = UserAccount(user_openid=openid, balance=balance)
            db.add(account)
            db.flush()
            logger.info(f"创建新账户 - openid: {openid}")

        return account.to_dict()


def get_balance(openid: str) -> dict:
    """查询用户余额"""
    with db_session() as db:
        account = (
            db.query(UserAccount)
            .filter(UserAccount.user_openid == openid, UserAccount.is_deleted == False)
            .first()
        )

        if not account:
            # 自动创建账户
            return get_or_create_account(openid)

        return account.to_dict()


def get_monthly_usage(openid: str, month: str) -> dict:
    """
    获取月度用量统计

    Args:
        openid: 用户openid
        month: 月份 "YYYY-MM"

    Returns:
        {
            "month": "2025-01",
            "summary": {
                "total_amount": 123.45,
                "total_api_calls": 150,
                "total_tokens": 50000
            },
            "daily_stats": [
                {
                    "date": "2025-01-01",
                    "amount": 10.50,
                    "api_calls": 12,
                    "tokens": 3500
                }
            ]
        }
    """
    try:
        # 解析月份
        start_date = datetime.strptime(f"{month}-01", "%Y-%m-%d")
        # 计算月末日期
        if start_date.month == 12:
            end_date = start_date.replace(year=start_date.year + 1, month=1, day=1)
        else:
            end_date = start_date.replace(month=start_date.month + 1, day=1)

        with db_session() as db:
            # 查询月度数据
            records = (
                db.query(TokenUsageRecord)
                .filter(
                    TokenUsageRecord.user_openid == openid,
                    TokenUsageRecord.is_deleted == False,
                    TokenUsageRecord.status == "success",
                    TokenUsageRecord.created_at >= start_date,
                    TokenUsageRecord.created_at < end_date,
                )
                .all()
            )

            # 按日期分组统计
            daily_map = {}
            total_amount = Decimal("0")
            total_calls = 0
            total_tokens = 0

            for record in records:
                date_key = record.created_at.strftime("%Y-%m-%d")

                if date_key not in daily_map:
                    daily_map[date_key] = {
                        "date": date_key,
                        "amount": Decimal("0"),
                        "api_calls": 0,
                        "tokens": 0,
                    }

                daily_map[date_key]["amount"] += record.cost or Decimal("0")
                daily_map[date_key]["api_calls"] += 1
                daily_map[date_key]["tokens"] += record.total_tokens or 0

                total_amount += record.cost or Decimal("0")
                total_calls += 1
                total_tokens += record.total_tokens or 0

            # 转换为列表并排序
            daily_stats = sorted(daily_map.values(), key=lambda x: x["date"])

            # Decimal转float
            for stat in daily_stats:
                stat["amount"] = float(stat["amount"])

            logger.info(
                f"查询月度用量成功 - user: {openid}, month: {month}, calls: {total_calls}"
            )

            return {
                "month": month,
                "summary": {
                    "total_amount": float(total_amount),
                    "total_api_calls": total_calls,
                    "total_tokens": total_tokens,
                },
                "daily_stats": daily_stats,
            }

    except ValueError as e:
        logger.error(f"月份格式错误 - month: {month}, error: {str(e)}")
        raise ValueError("月份格式错误，请使用 YYYY-MM 格式")
    except Exception as e:
        logger.error(f"查询月度用量失败 - error: {str(e)}")
        raise


def export_monthly_usage(openid: str, month: str) -> bytes:
    """
    导出月度用量Excel

    Args:
        openid: 用户openid
        month: 月份 "YYYY-MM"

    Returns:
        Excel文件字节流
    """

    # 获取用量数据
    usage_data = get_monthly_usage(openid, month)

    # 创建工作簿
    wb = Workbook()
    ws = wb.active
    ws.title = "用量报表"

    # 设置表头
    headers = ["日期", "扣费金额(元)", "API调用次数", "Token消耗量"]
    ws.append(headers)

    # 表头样式
    header_fill = PatternFill(
        start_color="4472C4", end_color="4472C4", fill_type="solid"
    )
    header_font = Font(bold=True, color="FFFFFF")

    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")

    # 填充数据
    for stat in usage_data["daily_stats"]:
        ws.append([stat["date"], stat["amount"], stat["api_calls"], stat["tokens"]])

    # 添加汇总行
    summary = usage_data["summary"]
    ws.append(
        [
            "合计",
            summary["total_amount"],
            summary["total_api_calls"],
            summary["total_tokens"],
        ]
    )

    # 汇总行样式
    summary_row = ws.max_row
    summary_fill = PatternFill(
        start_color="FFC000", end_color="FFC000", fill_type="solid"
    )
    summary_font = Font(bold=True)

    for cell in ws[summary_row]:
        cell.fill = summary_fill
        cell.font = summary_font
        cell.alignment = Alignment(horizontal="center", vertical="center")

    # 调整列宽
    ws.column_dimensions["A"].width = 15
    ws.column_dimensions["B"].width = 18
    ws.column_dimensions["C"].width = 18
    ws.column_dimensions["D"].width = 18

    # 保存到字节流
    output = BytesIO()
    wb.save(output)
    output.seek(0)

    logger.info(f"导出月度用量成功 - user: {openid}, month: {month}")

    return output.getvalue()

