import xlrd
from datetime import datetime
from pathlib import Path
from openpyxl import load_workbook
from app.utils.logger import get_logger

logger = get_logger(__name__)


def parse_excel(filepath):
    """解析 Excel 文件（支持 .xls 和 .xlsx）"""
    try:
        file_ext = Path(filepath).suffix.lower()

        if file_ext == ".xls":
            # 用 xlrd 解析 .xls
            return _parse_xls(filepath)
        else:
            # 用 openpyxl 解析 .xlsx
            return _parse_xlsx(filepath)

    except Exception as e:
        error_msg = f"Excel 解析失败: {str(e)}"
        logger.error(f"{error_msg} - 文件: {filepath}")
        raise Exception(error_msg)


def _parse_xls(filepath: str) -> str:
    """解析 .xls 文件"""
    xls_book = xlrd.open_workbook(filepath)

    if xls_book.nsheets == 0:
        logger.warning(f"Excel 文件为空: {filepath}")
        return "[Excel 文件无内容]"

    sheets_content = []

    for sheet_index in range(xls_book.nsheets):
        sheet = xls_book.sheet_by_index(sheet_index)
        rows = []

        for row_idx in range(sheet.nrows):
            row_values = []
            for col_idx in range(sheet.ncols):
                cell_value = sheet.cell_value(row_idx, col_idx)
                cell_type = sheet.cell_type(row_idx, col_idx)

                # 处理不同类型的单元格
                if cell_type == 3:  # 日期类型
                    try:
                        date_value = xlrd.xldate_as_datetime(
                            cell_value, xls_book.datemode
                        )
                        row_values.append(str(date_value))
                    except Exception:
                        row_values.append(str(cell_value))
                elif cell_value:  # 非空值
                    row_values.append(str(cell_value))

            # 拼接行内容
            row_text = " ".join(row_values)
            if row_text.strip():
                rows.append(row_text)

        if rows:
            sheets_content.append("\n".join(rows))

    content = "\n\n".join(sheets_content)

    logger.info(f"Excel 解析成功 (.xls) - 文件: {filepath}, 字符数: {len(content)}")

    return content if content.strip() else "[Excel 未识别到文字内容]"


def _parse_xlsx(filepath: str) -> str:
    """解析 .xlsx 文件"""
    wb = load_workbook(filepath, data_only=True)

    if not wb.sheetnames:
        logger.warning(f"Excel 文件为空: {filepath}")
        return "[Excel 文件无内容]"

    sheets_content = []

    for sheet_name in wb.sheetnames:
        sheet = wb[sheet_name]
        rows = []

        for row in sheet.iter_rows(values_only=True):
            row_text = " ".join([str(cell) for cell in row if cell is not None])
            if row_text.strip():
                rows.append(row_text)

        if rows:
            sheets_content.append("\n".join(rows))

    content = "\n\n".join(sheets_content)

    logger.info(f"Excel 解析成功 (.xlsx) - 文件: {filepath}, 字符数: {len(content)}")

    return content if content.strip() else "[Excel 未识别到文字内容]"
